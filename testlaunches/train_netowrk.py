import datetime
import time

from project.networks.losses import MyMSE, InlierRatio, MaxDeviation, MeanDeviation
from project.trainer import train
from project import utils
import testlaunches.build_tables
import testlaunches.metaparams
import tensorflow as tf

from openpyxl import Workbook


def load_tables(folder: str, table_name: str):
    table = utils.import_csv_table(f"./solution_tables/{folder}/{table_name}.csv")
    table = utils.shuffle_table(table)
    return utils.split_table_by_inp(table)


train_params = testlaunches.metaparams.prepare_params()

names = []
for func in testlaunches.build_tables.list_sol_functions:
    names.append(func.__name__)

for func in testlaunches.build_tables.list_table_functions:
    names.append(func.__name__)

data = dict()
for distribution in testlaunches.metaparams.folders:
    data[distribution] = (dict())
    for func_name in names:
        x, y = load_tables(distribution, func_name)
        data[distribution][func_name] = (x, y)

validation_tables = {}
for func_name in names:
    x, y = load_tables('validation_data', func_name)
    validation_tables[func_name] = (x, y)

column_names = []
for i in range(26):
    column_names.append(chr(ord('A') + i))

for i in range(26):
    column_names.append('A' + chr(ord('A') + i))

for i in range(26):
    column_names.append('B' + chr(ord('A') + i))

for i in range(26):
    column_names.append('С' + chr(ord('A') + i))

shape_separater = '-'
path = ""
print("Start experiments")
for params in train_params:
    normalize = params["normalize"]
    loss_func = params["loss_func"]
    optimizer = params["optimizer"]
    validation_split = params["validation_split"]
    distribution = params["distribution"]
    epochs = params["epochs"]
    rate = params["rate"]
    metrics = params["metrics"]
    validation_metrics = params["validation_metrics"]

    wb = Workbook()

    optimizer_name = testlaunches.metaparams.get_optimizer_name(optimizer)
    sheet_name = f"{normalize} {optimizer_name}"
    validate_name = "validate " + sheet_name
    wb.create_sheet(sheet_name)
    wb.create_sheet(validate_name)
    column_offset = 0
    column_out_offset = 0
    row_offset = 1
    row_validate_offset = 1
    row_per_metric = 1

    loss_func_name = testlaunches.metaparams.get_loss_func_name(loss_func)

    print(f"Params. Loss func {loss_func_name}. Normalize + Optimizer {sheet_name}")
    for epoch in epochs:
        for eps in rate:

            wb[sheet_name][f'{column_names[column_out_offset + 1]}{row_offset}'] = f"loss = {loss_func_name}, " \
                                                                                   f"epochs = {epoch}, " \
                                                                                   f"rate = {eps}, " \
                                                                                   f"validation split = {validation_split}, " \
                                                                                   f"normalize = {normalize}"

            wb[sheet_name][f'{column_names[column_offset]}{row_offset + 1}'] = 'loss = MSE'

            wb[validate_name][
                f'{column_names[column_out_offset + 1]}{row_offset}'] = f"loss = {loss_func_name}, " \
                                                                        f"epochs = {epoch}, " \
                                                                        f"rate = {eps}, " \
                                                                        f"validation split = {validation_split}, " \
                                                                        f"normalize = {normalize}"

            wb[validate_name][f'{column_names[column_offset]}{row_offset + 1}'] = 'loss = MSE'

            for func_name in names:
                x, y = data[distribution][func_name]
                x_val, y_val = validation_tables[func_name]
                train_start_time = time.time()
                nets, history = train(x, y, eps=eps, epochs=epoch, validation_split=validation_split,
                                      normalize=normalize, use_rand_net=True,
                                      name_salt=func_name + shape_separater,
                                      loss_func=loss_func,
                                      optimizer=optimizer, metrics=metrics, experiments=True)

                # title, networks and 2 empty
                row_per_metric = 1 + len(nets) + 2

                for net_num, net_history in enumerate(history):
                    net_shape = f"{net_history.model.name[net_history.model.name.rfind(shape_separater):]}"
                    for metric_num, train_metrics in enumerate(net_history.history):
                        wb[sheet_name][
                            f'{column_names[column_out_offset]}{row_offset + 1 + metric_num * row_per_metric}'] = f'metric = {train_metrics}'
                        wb[sheet_name][
                            f'{column_names[column_out_offset + column_offset + 1]}{row_offset + 1 + metric_num * row_per_metric}'] = f'{func_name}'
                        wb[sheet_name][
                            f'{column_names[column_out_offset]}{row_offset + net_num + 1 + metric_num * row_per_metric + 1}'] = f'{net_shape[2:]}'
                        wb[sheet_name][
                            f'{column_names[column_out_offset + column_offset + 1]}{row_offset + net_num + 1 + metric_num * row_per_metric + 1}'] = f'{net_history.history[train_metrics][-1]}'
                train_end_time = time.time()
                validation_start_time = time.time()
                for net_num, net in enumerate(nets):
                    net_shape = "_".join(map(str, net.get_shape))
                    for metric_num, val_metrics in enumerate(validation_metrics):
                        loss_val = validation_metrics[val_metrics](tf.convert_to_tensor(y_val, dtype=tf.float32),
                                                                   net.feedforward(x_val))
                        wb[validate_name][
                            f'{column_names[column_out_offset]}{row_validate_offset + 1 + metric_num * row_per_metric}'] = f'metric = {val_metrics}'
                        wb[validate_name][
                            f'{column_names[column_out_offset + column_offset + 1]}{row_validate_offset + 1 + metric_num * row_per_metric}'] = f'{func_name}'
                        wb[validate_name][
                            f'{column_names[column_out_offset]}{row_validate_offset + net_num + 1 + metric_num * row_per_metric + 1}'] = f'{net_shape}'
                        wb[validate_name][
                            f'{column_names[column_out_offset + column_offset + 1]}{row_validate_offset + net_num + 1 + metric_num * row_per_metric + 1}'] = f'{loss_val}'
                validation_end_time = time.time()
                print(f"Epochs size = {epoch}. "
                      f"Train time = {train_end_time - train_start_time}. "
                      f"Validation time = {validation_end_time - validation_start_time}")
                column_offset += 1
            column_out_offset += column_offset + 2
            column_offset = 0

        row_offset += row_per_metric * (len(metrics) + 1) * 2 + 2
        row_validate_offset += row_per_metric * len(validation_metrics) + 2
        column_out_offset = 0
        column_offset = 0

        print("End epochs = ", epoch)
    date = datetime.datetime.now()
    date_str = date.strftime('%I_%M_%S%p on %B %d, %Y')
    wb.save(f"{path}{distribution}_{date_str}.xlsx")
    wb.close()
